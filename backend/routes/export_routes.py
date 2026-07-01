import json
import datetime
from io import BytesIO
from flask import Blueprint, request, send_file, jsonify
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from backend.db import get_db
from backend.routes.admin_routes import admin_required

export_bp = Blueprint("export", __name__, url_prefix="/api/admin/export")

# ── Colours ──────────────────────────────────────────────────────────────────
NAVY   = "002B5C"
ORANGE = "EA580C"
HEADER_BG = "003D82"
ALT_ROW    = "EEF4FF"
WHITE      = "FFFFFF"
GREEN_BG   = "DCFCE7"
RED_BG     = "FEF2F2"
AMBER_BG   = "FFFBEB"

def _border(style="thin"):
    s = Side(style=style, color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_font():
    return Font(name="Calibri", bold=True, color=WHITE, size=10)

def _cell_font(bold=False):
    return Font(name="Calibri", bold=bold, size=9, color="0F172A")

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _parse_date_range():
    """Returns (date_from, date_to) as datetime objects, or error string."""
    now = datetime.datetime.now()
    range_key = request.args.get("range", "today")

    if range_key == "today":
        d_from = now.replace(hour=0, minute=0, second=0, microsecond=0)
        d_to   = now.replace(hour=23, minute=59, second=59)

    elif range_key == "7days":
        d_from = (now - datetime.timedelta(days=6)).replace(hour=0, minute=0, second=0, microsecond=0)
        d_to   = now.replace(hour=23, minute=59, second=59)

    elif range_key == "month":
        d_from = (now - datetime.timedelta(days=29)).replace(hour=0, minute=0, second=0, microsecond=0)
        d_to   = now.replace(hour=23, minute=59, second=59)

    elif range_key == "custom":
        try:
            d_from = datetime.datetime.strptime(request.args["from"], "%Y-%m-%d")
            d_to   = datetime.datetime.strptime(request.args["to"],   "%Y-%m-%d").replace(
                hour=23, minute=59, second=59)
        except (KeyError, ValueError):
            return None, None, "Custom range requires ?from=YYYY-MM-DD&to=YYYY-MM-DD"

        # Enforce 3-month cap
        if (d_to - d_from).days > 92:
            return None, None, "Custom range cannot exceed 3 months"

    else:
        return None, None, "Invalid range"

    return d_from, d_to, None


def _status(renewal_dates_json, valid_until_str):
    now = datetime.datetime.now()
    today = now.strftime("%Y-%m-%d")
    dates = json.loads(renewal_dates_json or "[]")
    latest = dates[-1][:10] if dates else ""
    try:
        vu = datetime.datetime.strptime(valid_until_str, "%Y-%m-%d %H:%M:%S")
        expired = vu < now
    except Exception:
        expired = False
    if latest == today:
        return "Active"
    if expired:
        return "Expired"
    return "Pending"


def _build_excel(d_from, d_to):
    """Build the Excel workbook and return (BytesIO, filename)."""
    conn = get_db()
    rows = conn.execute("""
        SELECT
            wp.id, wp.permit_no, wp.work_order_no,
            wo.description     AS job_description,
            wo.order_type_desc AS order_type,
            wp.permit_subtype, wp.shift, wp.exact_location,
            wp.location_lat, wp.location_lng,
            wp.partner_no, wp.partner_name, wp.num_workmen,
            wp.renewal_dates, wp.created_by, wp.created_at, wp.valid_until
        FROM work_permits wp
        LEFT JOIN work_orders wo ON wp.work_order_no = wo.order_no
        WHERE wp.created_at BETWEEN ? AND ?
        ORDER BY wp.created_at DESC
    """, (d_from.strftime("%Y-%m-%d %H:%M:%S"), d_to.strftime("%Y-%m-%d %H:%M:%S"))).fetchall()

    jsa_map = {r["permit_id"]: r["doc_no"] for r in conn.execute("SELECT permit_id, doc_no FROM jsa_records")}
    ei_map  = {r["permit_id"]: r["iso_no"]
               for r in conn.execute("SELECT permit_id, iso_no FROM electrical_isolations WHERE iso_no IS NOT NULL")}
    conn.close()

    HEADERS = [
        ("#",               4),  ("Permit No",       16), ("Work Order No",   16),
        ("Job Description", 34), ("Order Type",      18), ("Permit Type",     13),
        ("Shift",           10), ("Exact Location",  28), ("Latitude",        13),
        ("Longitude",       13), ("Partner No",      13), ("Partner Name",    22),
        ("Workmen",         10), ("JSA Doc No",      18), ("EI ISO No",       18),
        ("Created By",      20), ("Created At",      18), ("Valid Until",     18),
        ("Status",          12),
    ]
    NCOLS    = len(HEADERS)
    LAST_COL = get_column_letter(NCOLS)

    range_key   = request.args.get("range", "today")
    range_label = range_key.upper()
    if range_label == "CUSTOM":
        range_label = f"{d_from.strftime('%d %b %Y')} – {d_to.strftime('%d %b %Y')}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Work Permits"

    # Title rows
    ws.merge_cells(f"A1:{LAST_COL}1")
    c = ws["A1"]
    c.value = "HPCL — Work Permit Register"
    c.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    c.fill = _fill(HEADER_BG); c.alignment = _center()
    ws.row_dimensions[1].height = 22

    ws.merge_cells(f"A2:{LAST_COL}2")
    c = ws["A2"]
    c.value = (f"Exported: {datetime.datetime.now().strftime('%d %b %Y %H:%M')}  |  "
               f"Period: {range_label}  |  Records: {len(rows)}")
    c.font = Font(name="Calibri", italic=True, size=8.5, color="BFD7FF")
    c.fill = _fill(NAVY); c.alignment = _center()
    ws.row_dimensions[2].height = 14

    # Column headers
    for ci, (label, width) in enumerate(HEADERS, start=1):
        c = ws.cell(row=3, column=ci, value=label)
        c.font = _header_font(); c.fill = _fill(HEADER_BG)
        c.alignment = _center(); c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[3].height = 18

    STATUS_FILL  = {"Active": GREEN_BG, "Pending": AMBER_BG, "Expired": RED_BG}
    STATUS_COLOR = {"Active": "166534",  "Pending": "92400E",  "Expired": "991B1B"}
    MONO_COLS = {2, 3, 14, 15}   # Permit No, WO No, JSA, EI
    CENTER_COLS = {1, 9, 10, 13} # #, lat, lng, workmen

    for row_num, p in enumerate(rows, start=1):
        er = row_num + 3
        alt = row_num % 2 == 0
        status = _status(p["renewal_dates"], p["valid_until"])
        values = [
            row_num, p["permit_no"], p["work_order_no"],
            p["job_description"] or "", p["order_type"] or "",
            p["permit_subtype"], p["shift"], p["exact_location"],
            p["location_lat"], p["location_lng"],
            p["partner_no"], p["partner_name"], p["num_workmen"],
            jsa_map.get(p["id"], ""), ei_map.get(p["id"], ""),
            p["created_by"], p["created_at"], p["valid_until"], status,
        ]
        for ci, val in enumerate(values, start=1):
            c = ws.cell(row=er, column=ci, value=val)
            c.border = _border()
            c.fill = _fill(ALT_ROW if alt else WHITE)
            if ci == 19:  # Status
                c.fill = _fill(STATUS_FILL.get(status, WHITE))
                c.font = Font(name="Calibri", bold=True, size=9, color=STATUS_COLOR.get(status, "0F172A"))
                c.alignment = _center()
            elif ci in MONO_COLS:
                c.font = Font(name="Courier New", bold=True, size=8.5, color=NAVY)
                c.alignment = _center()
            elif ci in (9, 10):
                c.font = _cell_font(); c.number_format = "0.000000"; c.alignment = _center()
            elif ci in CENTER_COLS:
                c.font = _cell_font(bold=(ci == 1)); c.alignment = _center()
            else:
                c.font = _cell_font(); c.alignment = _left()
        ws.row_dimensions[er].height = 15

    ws.freeze_panes = "D4"
    ws.auto_filter.ref = f"A3:{LAST_COL}{3 + len(rows)}"

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    filename = f"HPCL_WorkPermits_{range_key}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return buf, filename


@export_bp.route("/work-permits")
@admin_required
def export_work_permits():
    d_from, d_to, err = _parse_date_range()
    if err:
        return jsonify({"error": err}), 400
    buf, filename = _build_excel(d_from, d_to)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)
